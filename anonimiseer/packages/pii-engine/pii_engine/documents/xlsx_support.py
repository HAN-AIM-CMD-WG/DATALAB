"""XLSX-extractie en -herbouw via openpyxl.

Elke cel met een string-waarde wordt één blok. Numerieke cellen en
formules laten we ongemoeid — dat zouden false positives worden en we
willen de betekenis van berekeningen niet kapot maken.
"""

from __future__ import annotations

import io
from typing import TYPE_CHECKING

from openpyxl import load_workbook

from ._common import apply_replacements_to_text, group_replacements_per_block

if TYPE_CHECKING:  # pragma: no cover
    from . import AcceptedReplacement, Block, ExtractResult

_BLOCK_SEPARATOR = "\n"


def _cell_id(sheet_idx: int, sheet_name: str, cell_ref: str) -> str:
    # sheet_idx maakt de id stabiel ook als twee sheets dezelfde naam
    # zouden hebben (technisch onmogelijk in Excel, maar defensief).
    safe_name = sheet_name.replace(":", "_")
    return f"s{sheet_idx}:{safe_name}!{cell_ref}"


def extract_xlsx(file_bytes: bytes) -> "ExtractResult":
    from . import Block, ExtractResult

    wb = load_workbook(io.BytesIO(file_bytes), data_only=False)
    texts: list[str] = []
    blocks: list[Block] = []
    cursor = 0

    for sheet_idx, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str):
                    continue
                if not value.strip():
                    continue
                block_id = _cell_id(sheet_idx, sheet_name, cell.coordinate)
                start = cursor
                end = start + len(value)
                blocks.append(Block(id=block_id, kind="sheet-cell", start=start, end=end))  # type: ignore[arg-type]
                texts.append(value)
                cursor = end + len(_BLOCK_SEPARATOR)

    flat_text = _BLOCK_SEPARATOR.join(texts)
    return ExtractResult(flat_text=flat_text, blocks=blocks)


def apply_xlsx(
    file_bytes: bytes,
    replacements: list["AcceptedReplacement"],
    blocks: list["Block"],
    footer_note: str | None = None,
) -> bytes:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=False)
    grouped = group_replacements_per_block(blocks, replacements)

    # Bouw een snelle index van block_id → cel.
    # Block id format: "s{idx}:{name}!{ref}" — parse terug.
    for block_id, block_replacements in grouped.items():
        if not block_replacements:
            continue
        try:
            sheet_part, cell_ref = block_id.split("!", 1)
            # sheet_part = "s{idx}:{name}"
            _, sheet_name = sheet_part.split(":", 1)
        except ValueError:
            continue
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        cell = ws[cell_ref]
        original = cell.value
        if not isinstance(original, str):
            continue
        new_text = apply_replacements_to_text(original, block_replacements)
        if new_text != original:
            cell.value = new_text

    if footer_note:
        # Voeg een aparte sheet "_Anonimiseer" toe met het watermerk.
        # We kiezen een aparte sheet ipv vervuiling van bestaande sheets,
        # zodat formules en lay-out niet kapot gaan.
        sheet_name = "_Anonimiseer"
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        meta = wb.create_sheet(sheet_name)
        meta.column_dimensions["A"].width = 100
        for idx, line in enumerate(footer_note.splitlines() or [footer_note], start=1):
            meta.cell(row=idx, column=1, value=line)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


__all__ = ["apply_xlsx", "extract_xlsx"]
