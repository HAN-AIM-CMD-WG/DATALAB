"""Regressie-tests voor de XLSX-pijplijn.

Dekt twee scenario's die we in het v0.1.0-onderzoek tegenkwamen:

1. **Merged cells**: openpyxl rapporteert ``MergedCell`` voor de
   "deelnemende" cellen van een ``A1:B1``-merge met ``value=None``.
   Onze ``isinstance(value, str)``-check filtert die al weg — deze
   test borgt dat dat zo blijft.
2. **Hyperlinks**: Excel-cellen met een ``mailto:`` of ``https://``
   hyperlink houden hun target ook nadat ``cell.value`` is gewijzigd.
   Voor pseudonimisatie moeten we die hyperlink-URL ook strippen
   anders staat het origineel als verborgen leak in het bestand.
"""

from __future__ import annotations

import io

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.hyperlink import Hyperlink

from pii_engine.documents import AcceptedReplacement
from pii_engine.documents.xlsx_support import apply_xlsx, extract_xlsx


def _wb_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestMergedCells:
    def _build(self) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Telefoon"
        ws["B1"] = "06-12345678"
        ws.merge_cells("B1:C1")
        return _wb_bytes(wb)

    def test_merged_cell_yields_single_block(self):
        data = self._build()
        result = extract_xlsx(data)
        # B1 heeft de waarde, C1 is een MergedCell met None — slechts één
        # tekst-block voor het telefoonnummer.
        text_blocks = [
            b for b in result.blocks if "06-12345678" in result.flat_text[b.start : b.end]
        ]
        assert len(text_blocks) == 1

    def test_merged_cell_replacement_clean(self):
        data = self._build()
        result = extract_xlsx(data)

        idx = result.flat_text.index("06-12345678")
        reps = [
            AcceptedReplacement(
                start=idx,
                end=idx + len("06-12345678"),
                replacement="NL_PHONE_NUMBER_1",
                original="06-12345678",
            )
        ]
        out_bytes = apply_xlsx(data, reps, result.blocks)

        out_wb = load_workbook(io.BytesIO(out_bytes))
        out_ws = out_wb.active
        assert out_ws["B1"].value == "NL_PHONE_NUMBER_1"


class TestHyperlinkStrip:
    def _build(self, email: str) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "E-mailadres"
        cell = ws.cell(row=1, column=2, value=email)
        cell.hyperlink = Hyperlink(ref=cell.coordinate, target=f"mailto:{email}")
        return _wb_bytes(wb)

    def test_hyperlink_target_stripped_after_replace(self):
        email = "josverstappen@example.nl"
        data = self._build(email)

        # Sanity: in de originele input zit zowel de display-text als de
        # mailto:-target.
        input_xml = io.BytesIO(data)
        in_wb = load_workbook(input_xml)
        assert in_wb.active["B1"].hyperlink is not None
        assert email in in_wb.active["B1"].hyperlink.target

        result = extract_xlsx(data)
        idx = result.flat_text.index(email)
        reps = [
            AcceptedReplacement(
                start=idx,
                end=idx + len(email),
                replacement="EMAIL_ADDRESS_1",
                original=email,
            )
        ]
        out_bytes = apply_xlsx(data, reps, result.blocks)

        out_wb = load_workbook(io.BytesIO(out_bytes))
        out_cell = out_wb.active["B1"]
        # Display-tekst vervangen door placeholder.
        assert out_cell.value == "EMAIL_ADDRESS_1"
        # En de hyperlink-target met daarin de originele email is weg.
        assert out_cell.hyperlink is None, (
            f"hyperlink-leak: target={out_cell.hyperlink.target!r}" if out_cell.hyperlink else "ok"
        )
        # Geen email-string te vinden in de gehele XLSX-bytes (in cel,
        # gedeelde strings, hyperlinks, enz.).
        assert email.encode() not in out_bytes, "originele email zit nog ergens in de XLSX-bytes"
